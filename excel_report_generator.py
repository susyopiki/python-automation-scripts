
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# --- Configure this ---
INPUT_CSV  = "sales_data.csv"   # Your raw data file
OUTPUT_FILE = "report.xlsx"     # The report it will generate

def generate_report(csv_file, output_file):
    # Read CSV data
    with open(csv_file, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    if not rows:
        print("No data found in CSV.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # --- Header styling ---
    header_font    = Font(bold=True, color="FFFFFF", size=12)
    header_fill    = PatternFill("solid", fgColor="2A52BE")
    center_align   = Alignment(horizontal="center")

    headers = list(rows[0].keys())

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header.upper())
        cell.font        = header_font
        cell.fill        = header_fill
        cell.alignment   = center_align

    # --- Write data rows ---
    for row_num, row in enumerate(rows, 2):
        for col_num, key in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row[key])
            cell.alignment = Alignment(horizontal="center")

            # Alternate row colour
            if row_num % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="E8EEFF")

    # --- Auto-fit column widths ---
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = max(
            len(str(header)),
            *[len(str(row[header])) for row in rows]
        )
        ws.column_dimensions[col_letter].width = max_length + 4

    # --- Summary row at the bottom ---
    ws.append([])
    summary_row = ws.max_row + 1

    for col_num, header in enumerate(headers, 1):
        values = []
        for row in rows:
            try:
                values.append(float(row[header]))
            except (ValueError, TypeError):
                pass
        if values:
            cell = ws.cell(row=summary_row, column=col_num, value=sum(values))
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="2A52BE")
            cell.font = Font(bold=True, color="FFFFFF")

    wb.save(output_file)
    print(f"Report saved as {output_file}")

generate_report(INPUT_CSV, OUTPUT_FILE)